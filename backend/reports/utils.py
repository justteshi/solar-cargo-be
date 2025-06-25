import io
import os
import subprocess
import threading
import time
from datetime import datetime
from copy import copy
from pathlib import Path
import logging
import math

import requests
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.contrib.auth import get_user_model
from django.conf import settings
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles.borders import Border, Side
from PIL import Image as PILImage

# Constants
ITEMS_PER_PAGE = 7
TICK = "✓"
CELL_MAP = {
    'location': 'A3',
    'checking_company': 'E3',
    'supplier': 'C9',
    'delivery_slip_number': 'C10',
    'logistic_company': 'C11',
    'container_number': 'C12',
    'licence_plate_truck': 'C13',
    'licence_plate_trailer': 'C14',
    'weather_conditions': 'C15',
    'user': 'D45',
}
STATUS_FIELDS = {
    'load_secured_status': 18,
    'delivery_without_damages_status': 19,
    'packaging_status': 20,
    'goods_according_status': 21,
    'suitable_machines_status': 22,
    'delivery_slip_status': 23,
    'inspection_report_status': 24,
}
COMMENT_FIELD_MAP = {
    'load_secured_status': 'load_secured_comment',
    'delivery_without_damages_status': 'delivery_without_damages_comment',
    'packaging_status': 'packaging_comment',
    'goods_according_status': 'goods_according_comment',
    'suitable_machines_status': 'suitable_machines_comment',
    'delivery_slip_status': 'delivery_slip_comment',
    'inspection_report_status': 'inspection_report_comment',
}
IMAGE_MAP_BASE_ROW = 29

# Lock to ensure only one API call at a time
_plate_recognizer_lock = threading.Lock()
logger = logging.getLogger(__name__)


class PlateRecognitionError(Exception):
    """Custom exception for plate recognition errors."""
    pass


def recognize_plate(image_path):
    """Recognize license plate from an image using Plate Recognizer API."""
    api_key = os.environ.get("PLATE_RECOGNIZER_API_KEY")
    if not api_key:
        raise RuntimeError("❌ Plate Recognizer API key is not set in environment variables.")

    with _plate_recognizer_lock:
        time.sleep(1)
        with open(image_path, 'rb') as img_file:
            response = requests.post(
                'https://api.platerecognizer.com/v1/plate-reader/',
                files={'upload': img_file},
                headers={'Authorization': f'Token {api_key}'}
            )
        if response.status_code not in (200, 201):
            raise PlateRecognitionError(f"❌ API request failed: {response.status_code} - {response.text}")

        data = response.json()
        try:
            plate = data['results'][0]['plate'].upper()
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print(f"✅ {timestamp} — License plate recognized successfully: {plate}")
            return plate
        except (KeyError, IndexError):
            raise PlateRecognitionError("❌ No license plate detected in the image.")


def save_report_to_excel(data, file_path=None, template_path='delivery_report_template.xlsx'):
    """
    Save delivery report data to an Excel file using a template.
    Returns the absolute path for further processing (e.g., PDF conversion).
    """
    relative_path, abs_path = get_relative_and_abs_path(file_path)
    items = data.get("items", [])
    extra_rows = max(0, len(items) - ITEMS_PER_PAGE)
    # Always load from template if file does not exist
    if default_storage.exists(relative_path):
        with default_storage.open(relative_path, 'rb') as f:
            wb = load_workbook(f)
    elif Path(abs_path).exists():
        with open(abs_path, 'rb') as f:
            wb = load_workbook(f)
    else:
        with open(template_path, 'rb') as f:
            wb = load_workbook(f)
        ws = wb.active
        # Write regular fields
        for key, cell in CELL_MAP.items():
            if key in data:
                value = data[key]
                if isinstance(value, bool):
                    value = "Yes" if value else "No"
                target_cell = get_top_left_cell(ws, cell)
                ws[target_cell] = value
                if key in ['location', 'checking_company']:
                    ws[target_cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                elif key == 'user':
                    ws[target_cell].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Write status fields and comments
    for field, row in STATUS_FIELDS.items():
        value = data.get(field, None)
        for col, cond in zip(['I', 'J', 'K'], [True, False, None]):
            cell = f'{col}{row}'
            target_cell = get_top_left_cell(ws, cell)
            ws[target_cell] = TICK if ((value is True and col == 'I') or (value is False and col == 'J') or (
                    value is None and col == 'K')) else ""
            ws[target_cell].alignment = Alignment(horizontal='center', vertical='center')
        comment_key = COMMENT_FIELD_MAP.get(field)
        if comment_key is not None and comment_key in data:
            comment_cell = f"L{row}"
            top_left_comment_cell = get_top_left_cell(ws, comment_cell)
            ws[top_left_comment_cell] = data[comment_key]
            ws[top_left_comment_cell].alignment = Alignment(wrap_text=True, vertical='top')
            # Only resize the correct cell(s)
            if extra_rows == 0:
                autofit_row_height(ws, top_left_comment_cell, data[comment_key], multiplier=25)
            else:
                offset_row = row + extra_rows
                offset_cell = f"L{offset_row}"
                top_left_offset_cell = get_top_left_cell(ws, offset_cell)
                autofit_row_height(ws, top_left_offset_cell, data[comment_key], multiplier=25)

   # Write items (first page only)
    write_items_to_excel(ws, items[:ITEMS_PER_PAGE], start_row=9)

    # Calculate extra_rows before placing images
    items = data.get("items", [])
    extra_rows = max(0, len(items) - ITEMS_PER_PAGE)

    # Dynamically adjust image section range
    image_start_row = 29 + extra_rows
    image_end_row = 42 + extra_rows
    image_start_cell = f"A{image_start_row}"
    image_end_cell = f"L{image_end_row}"

    # Collect image URLs as before
    image_urls = [
        data.get('truck_license_plate_image'),
        data.get('trailer_license_plate_image'),
        data.get('proof_of_delivery_image'),
    ]
    image_urls += [img.get('image') for img in data.get('additional_images_urls', []) if img.get('image')]
    image_urls = [url for url in image_urls if url]

    # Place images in the dynamically calculated range
    create_collage_of_images(ws, image_urls, image_start_cell, image_end_cell)
    # --- End refactored image section ---

    # Date
    ws['D46'] = datetime.now().strftime("%Y-%m-%d")
    ws['D46'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Save to in-memory buffer and then to storage
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    default_storage.save(relative_path, ContentFile(output.read()))

    # Append extra items if needed
    if len(items) > ITEMS_PER_PAGE:
        append_extra_items_to_excel(relative_path, items[ITEMS_PER_PAGE:], 9 + ITEMS_PER_PAGE)
        # Now reload and write comments at the correct row
        with default_storage.open(relative_path, 'rb') as f:
            wb = load_workbook(f)
            ws = wb.active
        extra_rows = len(items) - ITEMS_PER_PAGE
    else:
        # No extra items, comments row is 27
        with default_storage.open(relative_path, 'rb') as f:
            wb = load_workbook(f)
            ws = wb.active
        extra_rows = 0

    # Write comments at the correct row
    comments_row = 27 + extra_rows
    comments_cell = f"A{comments_row}"
    if 'comments' in data:
        ws[comments_cell] = data['comments']
        ws[comments_cell].alignment = Alignment(wrap_text=True, vertical='top')
        autofit_row_height(ws, comments_cell, data['comments'], multiplier=1.5)

    cmr_url = data.get('cmr_image')
    delivery_slip_url = data.get('delivery_slip_image')
    if cmr_url or delivery_slip_url:
        insert_cmr_delivery_slip_images(ws, cmr_url, delivery_slip_url)

    # Save again
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    with default_storage.open(relative_path, 'wb') as f:
        f.write(output.read())

    return abs_path

def write_items_to_excel(ws, items, start_row):
    """
    Write delivery items to the Excel worksheet starting at start_row.
    """
    for idx, entry in enumerate(items):
        row = start_row + idx
        ws[get_top_left_cell(ws, f'E{row}')] = "Item"
        ws[get_top_left_cell(ws, f'F{row}')] = entry["item"]["name"]
        ws[get_top_left_cell(ws, f'I{row}')] = "Amount"
        ws[get_top_left_cell(ws, f'J{row}')] = entry["quantity"]


def append_extra_items_to_excel(file_path, extra_items, insert_row):
    """
    Append extra delivery items to an existing Excel file, preserving formatting.
    """
    with default_storage.open(file_path, 'rb') as f:
        wb = load_workbook(f)
        ws = wb.active

        # Store and unmerge affected merged cells
        affected_merges = []
        for rng in list(ws.merged_cells.ranges):
            if rng.min_row >= insert_row:
                affected_merges.append((rng.min_row, rng.min_col, rng.max_row, rng.max_col))
                ws.unmerge_cells(str(rng))

        # Insert rows and copy style
        ws.insert_rows(insert_row, amount=len(extra_items))
        row_above = insert_row - 1
        for i in range(len(extra_items)):
            copy_row_style(ws, row_above, insert_row + i, min_col=1, max_col=12)

        # Merge and align cells
        for i in range(len(extra_items)):
            row = insert_row + i
            ws.merge_cells(f"A{row}:B{row}")
            ws.merge_cells(f"C{row}:D{row}")
            ws.merge_cells(f"F{row}:H{row}")
            ws.merge_cells(f"J{row}:L{row}")
            # Set right border for column L
            cell = ws.cell(row=row, column=12)
            original = cell.border
            cell.border = Border(
                left=original.left,
                right=Side(style="thick"),
                top=original.top,
                bottom=original.bottom,
                diagonal=original.diagonal,
                diagonal_direction=original.diagonal_direction,
                outline=original.outline,
                vertical=original.vertical,
                horizontal=original.horizontal,
            )

        # Re-merge cells in new positions
        for min_row, min_col, max_row, max_col in affected_merges:
            new_min_row = min_row + len(extra_items)
            new_max_row = max_row + len(extra_items)
            start_cell = f"{get_column_letter(min_col)}{new_min_row}"
            end_cell = f"{get_column_letter(max_col)}{new_max_row}"
            ws.merge_cells(f"{start_cell}:{end_cell}")

        # Write extra items
        write_items_to_excel(ws, extra_items, insert_row)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        with default_storage.open(file_path, 'wb') as f:
            f.write(output.read())


def get_top_left_cell(ws, cell):
    """
    Return the top-left cell of a merged range if cell is merged, else itself.
    """
    for merged_range in ws.merged_cells.ranges:
        if cell in merged_range:
            return merged_range.start_cell.coordinate
    return cell


def get_range_dimensions(ws, start_cell, end_cell):
    """
    Calculate pixel width and height of a cell range.
    """
    start_col = ''.join(filter(str.isalpha, start_cell))
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    end_col = ''.join(filter(str.isalpha, end_cell))
    end_row = int(''.join(filter(str.isdigit, end_cell)))

    total_width = 0
    for col_idx in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1):
        col_letter = get_column_letter(col_idx)
        total_width += (ws.column_dimensions[col_letter].width or 8.43) * 7

    total_height = 0
    for row in range(start_row, end_row + 1):
        total_height += (ws.row_dimensions[row].height or 15) * 0.75
    return int(total_width), int(total_height)


def copy_row_style(ws, src_row, dest_row, min_col=1, max_col=12):
    """
    Copy cell styles from src_row to dest_row for specified columns.
    """
    for col in range(min_col, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dest_cell = ws.cell(row=dest_row, column=col)
        if src_cell.has_style:
            dest_cell.font = copy(src_cell.font)
            dest_cell.border = copy(src_cell.border)
            dest_cell.fill = copy(src_cell.fill)
            dest_cell.number_format = copy(src_cell.number_format)
            dest_cell.protection = copy(src_cell.protection)
            dest_cell.alignment = copy(src_cell.alignment)


def get_username_from_id(user_id):
    """
    Return the full name for a user ID, or username if full name is blank.
    """
    User = get_user_model()
    try:
        user = User.objects.get(pk=user_id)
        full_name = user.get_full_name()
        return full_name if full_name.strip() else user.username
    except User.DoesNotExist:
        return "Unknown User"


def convert_excel_to_pdf(excel_path):
    """
    Convert an Excel file to PDF using LibreOffice.
    Returns the path to the generated PDF.
    """
    base_dir = Path(excel_path).parent
    pdf_dir = base_dir / "PDF files"
    pdf_dir.mkdir(exist_ok=True)

    command = [
        'libreoffice',
        '--headless',
        '--convert-to', 'pdf',
        '--outdir', str(pdf_dir),
        str(excel_path)
    ]
    logger.debug(f"Running command: {' '.join(command)}")
    result = subprocess.run(command, capture_output=True)
    logger.debug(f"LibreOffice return code: {result.returncode}")
    logger.debug(f"LibreOffice stdout: {result.stdout.decode()}")
    logger.debug(f"LibreOffice stderr: {result.stderr.decode()}")

    if result.returncode != 0:
        raise RuntimeError(f"LibreOffice failed: {result.stderr.decode()}")

    pdf_filename = Path(excel_path).with_suffix('.pdf').name
    pdf_path = pdf_dir / pdf_filename
    logger.debug(f"Expected PDF path: {pdf_path}")
    logger.debug(f"PDF file exists: {pdf_path.exists()}")
    return str(pdf_path)

def autofit_row_height(ws, cell, text, multiplier=14):
    """
    Set row height so all text is visible, minimizing excess whitespace.
    """
    col_letter = ''.join(filter(str.isalpha, cell))
    row_num = int(''.join(filter(str.isdigit, cell)))
    col_width = ws.column_dimensions[col_letter].width or 8.43
    chars_per_line = int(col_width * 1.15)
    text = (text or "").rstrip()
    total_lines = 0
    for line in text.split('\n'):
        line = line.rstrip()
        if not line:
            total_lines += 1
        else:
            total_lines += (len(line) - 1) // chars_per_line + 1
    ws.row_dimensions[row_num].height = max(15, total_lines * multiplier)

def get_relative_and_abs_path(file_path=None, subdir="delivery_reports", ext="xlsx"):
    """
    Returns (relative_path, abs_path) for storage and local use.
    """
    if file_path:
        abs_path = Path(file_path)
        relative_path = abs_path.relative_to(settings.MEDIA_ROOT)
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        relative_path = Path(subdir) / f"delivery_report_{timestamp}.{ext}"
        abs_path = Path(settings.MEDIA_ROOT) / relative_path
    abs_path.parent.mkdir(parents=True, exist_ok=True)
    return str(relative_path), str(abs_path)

def create_collage_of_images(ws, image_urls, start_cell, end_cell):
    n_images = len(image_urls)
    if n_images == 0:
        return

    max_width, max_height = get_range_dimensions(ws, start_cell, end_cell)
    area_aspect_ratio = max_width / max_height

    # Find the best grid layout
    best_layout = None
    best_size = 0
    for cols in range(1, n_images + 1):
        rows = math.ceil(n_images / cols)
        cell_width = max_width // cols
        cell_height = max_height // rows
        size = min(cell_width, cell_height)
        grid_aspect_ratio = (cols * cell_width) / (rows * cell_height)
        if size > best_size or (size == best_size and abs(grid_aspect_ratio - area_aspect_ratio) < abs((best_layout or (0,0,0))[2] - area_aspect_ratio)):
            best_layout = (cols, rows, grid_aspect_ratio)
            best_size = size

    cols, rows, _ = best_layout
    cell_width = max_width // cols
    cell_height = max_height // rows

    # Calculate actual used width and height
    last_row_images = n_images % cols if n_images % cols != 0 else cols
    used_width = cell_width * (cols if n_images > cols else n_images) if n_images > (rows - 1) * cols else cell_width * last_row_images
    used_height = cell_height * (rows - 1) + (cell_height if last_row_images else 0)

    collage = PILImage.new("RGBA", (cell_width * cols, cell_height * rows), (255, 255, 255, 0))

    for idx, url in enumerate(image_urls):
        try:
            response = requests.get(url)
            response.raise_for_status()
            img_bytes = io.BytesIO(response.content)
            pil_img = PILImage.open(img_bytes)
            pil_img.thumbnail((cell_width, cell_height), PILImage.LANCZOS)
            x = (idx % cols) * cell_width + (cell_width - pil_img.width) // 2
            y = (idx // cols) * cell_height + (cell_height - pil_img.height) // 2
            collage.paste(pil_img, (x, y))
        except Exception as e:
            logger.error(f"Error adding image to collage: {e}")

    # Crop to actual used area (removes trailing whitespace and prevents overflow)
    collage = collage.crop((0, 0, used_width, used_height))

    output_img = io.BytesIO()
    collage.save(output_img, format='PNG')
    output_img.seek(0)
    img = XLImage(output_img)
    img.anchor = start_cell
    ws.add_image(img)

def insert_cmr_delivery_slip_images(ws, cmr_url=None, delivery_slip_url=None):
    """
    Insert CMR and delivery slip images as full-size images, each on a new worksheet.
    """
    from openpyxl import Workbook

    wb = ws.parent  # Get the workbook from the current worksheet
    image_data = [
        ("CMR", cmr_url),
        ("Delivery Slip", delivery_slip_url)
    ]

    for sheet_name, url in image_data:
        if not url:
            continue
        try:
            # Create a new worksheet for each image
            img_ws = wb.create_sheet(title=sheet_name)
            response = requests.get(url)
            response.raise_for_status()
            img_bytes = io.BytesIO(response.content)
            pil_img = PILImage.open(img_bytes)
            output_img = io.BytesIO()
            pil_img.save(output_img, format='PNG')
            output_img.seek(0)
            xl_img = XLImage(output_img)
            xl_img.anchor = "A1"
            img_ws.add_image(xl_img)
            # Optionally, set row height/column width for better fit
            img_ws.row_dimensions[1].height = pil_img.height * 0.75 / 1.33
            img_ws.column_dimensions['A'].width = pil_img.width / 7
        except Exception as e:
            logger.error(f"Error inserting full-size image on new sheet: {e}")