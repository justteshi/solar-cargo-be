import io
import math
import requests
import logging
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as XLImage

logger = logging.getLogger(__name__)

def create_collage_of_images(ws, image_urls, start_cell, end_cell):
    n_images = len(image_urls)
    if n_images == 0:
        return
    max_width, max_height = get_range_dimensions(ws, start_cell, end_cell)
    area_aspect_ratio = max_width / max_height
    best_layout = None
    best_size = 0
    for cols in range(1, n_images + 1):
        rows = math.ceil(n_images / cols)
        cell_width = max_width // cols
        cell_height = max_height // rows
        size = min(cell_width, cell_height)
        grid_aspect_ratio = (cols * cell_width) / (rows * cell_height)
        if size > best_size or (size == best_size and abs(grid_aspect_ratio - area_aspect_ratio) < abs((best_layout or (0,0,0))[2] - area_aspect_ratio)):
            best_layout = (cols, rows, grid_aspect_ratio)
            best_size = size
    cols, rows, _ = best_layout
    cell_width = max_width // cols
    cell_height = max_height // rows
    last_row_images = n_images % cols if n_images % cols != 0 else cols
    used_width = cell_width * (cols if n_images > cols else n_images) if n_images > (rows - 1) * cols else cell_width * last_row_images
    used_height = cell_height * (rows - 1) + (cell_height if last_row_images else 0)
    collage = PILImage.new("RGBA", (cell_width * cols, cell_height * rows), (255, 255, 255, 0))
    for idx, url in enumerate(image_urls):
        try:
            response = requests.get(url)
            response.raise_for_status()
            img_bytes = io.BytesIO(response.content)
            pil_img = PILImage.open(img_bytes)
            pil_img.thumbnail((cell_width, cell_height), PILImage.LANCZOS)
            x = (idx % cols) * cell_width + (cell_width - pil_img.width) // 2
            y = (idx // cols) * cell_height + (cell_height - pil_img.height) // 2
            collage.paste(pil_img, (x, y))
        except Exception as e:
            logger.error(f"Error adding image to collage: {e}")
    collage = collage.crop((0, 0, used_width, used_height))
    output_img = io.BytesIO()
    collage.save(output_img, format='PNG')
    output_img.seek(0)
    img = XLImage(output_img)
    img.anchor = start_cell
    ws.add_image(img)

def insert_cmr_delivery_slip_images(ws, cmr_url=None, delivery_slip_url=None):
    wb = ws.parent
    image_data = [
        ("CMR", cmr_url),
        ("Delivery Slip", delivery_slip_url)
    ]
    for sheet_name, url in image_data:
        if not url:
            continue
        try:
            img_ws = wb.create_sheet(title=sheet_name)
            response = requests.get(url)
            response.raise_for_status()
            img_bytes = io.BytesIO(response.content)
            pil_img = PILImage.open(img_bytes)
            output_img = io.BytesIO()
            pil_img.save(output_img, format='PNG')
            output_img.seek(0)
            xl_img = XLImage(output_img)
            xl_img.anchor = "A1"
            img_ws.add_image(xl_img)
            img_ws.row_dimensions[1].height = pil_img.height * 0.75 / 1.33
            img_ws.column_dimensions['A'].width = pil_img.width / 7
        except Exception as e:
            logger.error(f"Error inserting full-size image on new sheet: {e}")

def get_range_dimensions(ws, start_cell, end_cell):
    from openpyxl.utils import get_column_letter, column_index_from_string
    start_col = ''.join(filter(str.isalpha, start_cell))
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    end_col = ''.join(filter(str.isalpha, end_cell))
    end_row = int(''.join(filter(str.isdigit, end_cell)))
    total_width = 0
    for col_idx in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1):
        col_letter = get_column_letter(col_idx)
        total_width += (ws.column_dimensions[col_letter].width or 8.43) * 7
    total_height = 0
    for row in range(start_row, end_row + 1):
        total_height += (ws.row_dimensions[row].height or 15) * 0.75
    return int(total_width), int(total_height)