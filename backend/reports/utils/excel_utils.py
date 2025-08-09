import io
from io import BytesIO
import logging
from datetime import datetime
from pathlib import Path
from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage, ImageOps
from openpyxl.worksheet.pagebreak import Break

from .image_utils import fetch_image_bytes, create_collage_of_images, insert_cmr_sheet, insert_images_in_single_sheet

logger = logging.getLogger(__name__)

ITEMS_PER_PAGE = 7
TICK = "✓"
CELL_MAP = {
    'location': 'A3',
    'supplier_name': 'C9',
    'delivery_slip_number': 'C10',
    'logistic_company': 'C11',
    'container_number': 'C12',
    'licence_plate_truck': 'C13',
    'licence_plate_trailer': 'C14',
    'weather_conditions': 'C15',
    'user': 'D44',
}
STATUS_FIELDS = {
    'load_secured_status': 18,
    'delivery_without_damages_status': 19,
    'packaging_status': 20,
    'goods_according_status': 21,
    'suitable_machines_status': 22,
    'delivery_slip_status': 23,
    'inspection_report_status': 24,
}
COMMENT_FIELD_MAP = {
    'load_secured_status': 'load_secured_comment',
    'delivery_without_damages_status': 'delivery_without_damages_comment',
    'packaging_status': 'packaging_comment',
    'goods_according_status': 'goods_according_comment',
    'suitable_machines_status': 'suitable_machines_comment',
    'delivery_slip_status': 'delivery_slip_comment',
    'inspection_report_status': 'inspection_report_comment',
}


def save_report_to_excel(data, file_path=None, template_path=None):
    if template_path is None:
        template_path = settings.REPORT_PATHS['TEMPLATE_PATH']
    relative_path, abs_path = get_relative_and_abs_path(
        file_path,
        subdir=settings.REPORT_PATHS['EXCEL_SUBDIR'],
        ext="xlsx"
    )
    items = data.get("items", [])
    extra_rows = max(0, len(items) - ITEMS_PER_PAGE)

    wb = None
    try:
        wb = _load_or_create_workbook(relative_path, abs_path, template_path, data)
        ws = wb.active

        # Handle different sections
        _handle_basic_fields(ws, data)
        _handle_client_logo(ws, data)
        _handle_client_name(ws, data)
        _handle_status_fields(ws, data, extra_rows)
        _handle_items_section(ws, items, extra_rows)
        _handle_image_sections(ws, data, extra_rows)
        _handle_date_field(ws)
        _handle_signature(ws, data, extra_rows)

        # Save initial workbook
        _save_workbook(wb, relative_path)

        # Handle extra items if needed
        if len(items) > ITEMS_PER_PAGE:
            append_extra_items_to_excel(relative_path, items[ITEMS_PER_PAGE:], 9 + ITEMS_PER_PAGE)

        # Final operations
        _handle_final_operations(relative_path, data, items)

        return abs_path

    finally:
        if wb:
            wb.close()


def _load_or_create_workbook(relative_path, abs_path, template_path, data):
    """Load existing workbook or create from template"""
    if default_storage.exists(relative_path):
        with default_storage.open(relative_path, 'rb') as f:
            return load_workbook(f)
    elif Path(abs_path).exists():
        with open(abs_path, 'rb') as f:
            return load_workbook(f)
    else:
        with open(template_path, 'rb') as f:
            wb = load_workbook(f)
        ws = wb.active
        _populate_basic_template_fields(ws, data)
        return wb


def _populate_basic_template_fields(ws, data):
    """Populate basic fields when creating from template"""
    for key, cell in CELL_MAP.items():
        if key in data and key != 'client_logo':
            value = data[key]
            if isinstance(value, bool):
                value = "Yes" if value else "No"
            target_cell = get_top_left_cell(ws, cell)
            ws[target_cell] = value
            _apply_cell_alignment(ws, target_cell, key)


def _apply_cell_alignment(ws, cell, key):
    """Apply specific alignment based on field type"""
    if key in ['location']:
        ws[cell].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    elif key == 'user':
        ws[cell].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _handle_basic_fields(ws, data):
    """Handle basic field population for existing workbooks"""
    for key, cell in CELL_MAP.items():
        if key in data and key != 'client_logo':
            value = data[key]
            if isinstance(value, bool):
                value = "Yes" if value else "No"
            target_cell = get_top_left_cell(ws, cell)
            ws[target_cell] = value
            _apply_cell_alignment(ws, target_cell, key)


def _handle_client_name(ws, data):
    """Insert client name"""
    client_name = data.get('location_client_name')
    if client_name:
        cell = ws['I3']
        cell.value = client_name
        cell.font = Font(size=12, name='Arial')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        border = Side(style="medium")
        cell.border = Border(left=cell.border.left, right=border, top=cell.border.top, bottom=cell.border.bottom)


def _handle_client_logo(ws, data):
    """Handle client logo insertion"""
    client_logo_url = data.get('client_logo')
    if client_logo_url:
        insert_client_logo(ws, client_logo_url, cell="I4", end_cell="L7")
    else:
        logger.info("No client logo URL provided.")


def _handle_status_fields(ws, data, extra_rows):
    """Handle status field checkboxes and comments"""
    for field, row in STATUS_FIELDS.items():
        _populate_status_checkboxes(ws, data, field, row)
        _populate_status_comments(ws, data, field, row, extra_rows)


def _populate_status_checkboxes(ws, data, field, row):
    """Populate status checkboxes (Yes/No/N/A)"""
    value = data.get(field, None)
    for col, cond in zip(['I', 'J', 'K'], [True, False, None]):
        cell = f'{col}{row}'
        target_cell = get_top_left_cell(ws, cell)
        ws[target_cell] = TICK if ((value is True and col == 'I') or
                                   (value is False and col == 'J') or
                                   (value is None and col == 'K')) else ""
        ws[target_cell].alignment = Alignment(horizontal='center', vertical='center')


def _populate_status_comments(ws, data, field, row, extra_rows):
    """Populate status field comments"""
    comment_key = COMMENT_FIELD_MAP.get(field)
    if comment_key is not None and comment_key in data:
        comment_cell = f"L{row}"
        top_left_comment_cell = get_top_left_cell(ws, comment_cell)
        ws[top_left_comment_cell] = data[comment_key]
        ws[top_left_comment_cell].alignment = Alignment(wrap_text=True, vertical='top')

        if extra_rows == 0:
            autofit_row_height(ws, top_left_comment_cell, data[comment_key], multiplier=15)
        else:
            offset_row = row + extra_rows
            offset_cell = f"L{offset_row}"
            top_left_offset_cell = get_top_left_cell(ws, offset_cell)
            autofit_row_height(ws, top_left_offset_cell, data[comment_key], multiplier=15)


def _handle_items_section(ws, items, extra_rows):
    """Handle items table population"""
    write_items_to_excel(ws, items[:ITEMS_PER_PAGE], start_row=9)


def _handle_image_sections(ws, data, extra_rows):
    """Handle all image insertions"""
    image_start_row = 28 + extra_rows
    image_end_row = 41 + extra_rows
    image_start_cell = f"A{image_start_row}"
    image_end_cell = f"L{image_end_row}"

    image_urls = [
        data.get('truck_license_plate_image'),
        data.get('trailer_license_plate_image'),
        data.get('proof_of_delivery_image'),
    ]
    image_urls = [url for url in image_urls if url]
    create_collage_of_images(ws, image_urls, image_start_cell, image_end_cell, row_offset=9)


def _handle_date_field(ws):
    """Set current date"""
    ws['J44'] = datetime.now().strftime("%Y-%m-%d")
    ws['J44'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _handle_signature(ws, data, extra_rows):
    """Insert user signature image, keeping height and stretching width to cell range."""
    signature_url = data.get('user_signature')
    if not signature_url:
        logger.info("No signature found for user.")
        return

    anchor_cell = f'G{43 + extra_rows}'
    start_col = column_index_from_string('G')
    start_row = 43 + extra_rows
    end_col = column_index_from_string('H')
    end_row = 45 + extra_rows

    total_width = sum(
        int((ws.column_dimensions[get_column_letter(col)].width or 8.43) * 7)
        for col in range(start_col, end_col + 1)
    )
    total_height = sum(
        int((ws.row_dimensions[row].height or 15) * 1.33)
        for row in range(start_row, end_row + 1)
    )

    img_bytes = BytesIO(fetch_image_bytes(signature_url))
    if not img_bytes.getbuffer().nbytes:
        logger.warning(f"Signature image could not be loaded from {signature_url}")
        return

    try:
        pil_img = PILImage.open(img_bytes)
        pil_img = ImageOps.exif_transpose(pil_img)
        pil_img = pil_img.convert("RGBA")

        # Resize width to total_width, keep original height (or limit to total_height)
        new_height = min(pil_img.height, total_height)
        pil_img = pil_img.resize((total_width, new_height), PILImage.LANCZOS)

        # Center vertically on transparent canvas if needed
        canvas = PILImage.new("RGBA", (total_width, total_height), (255, 255, 255, 0))
        y = (total_height - new_height) // 2
        canvas.paste(pil_img, (0, y), pil_img)

        output_img = BytesIO()
        canvas.save(output_img, format="PNG")
        output_img.seek(0)

        xl_img = XLImage(output_img)
        xl_img.anchor = anchor_cell
        ws.add_image(xl_img)
    except Exception as e:
        logger.error(f"Failed to insert signature image: {e}")


def _save_workbook(wb, relative_path):
    """Save workbook to storage"""
    output = None
    try:
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        default_storage.save(relative_path, ContentFile(output.read()))
    finally:
        if output:
            output.close()


def _handle_final_operations(relative_path, data, items):
    """Handle comments, damages, and additional sheets"""
    wb_final = None
    try:
        with default_storage.open(relative_path, 'rb') as f:
            wb_final = load_workbook(f)
            ws = wb_final.active

        extra_rows = len(items) - ITEMS_PER_PAGE if len(items) > ITEMS_PER_PAGE else 0

        _handle_comments_section(ws, data, extra_rows)
        _handle_page_break(ws)
        _handle_damages_section(ws, data)
        _handle_additional_sheets(wb_final, data)

        _save_final_workbook(wb_final, relative_path)

    finally:
        if wb_final:
            wb_final.close()


def _handle_comments_section(ws, data, extra_rows):
    """Handle comments section"""
    comments_row = 26 + extra_rows
    comments_cell = f"A{comments_row}"
    if 'comments' in data:
        ws[comments_cell] = data['comments']
        ws[comments_cell].alignment = Alignment(wrap_text=True, vertical='top')
        autofit_row_height(ws, comments_cell, data['comments'], multiplier=1.5)


def _handle_page_break(ws):
    """Insert manual page break"""
    ws.row_breaks.append(Break(id=ws.max_row + 1))


def _handle_damages_section(ws, data):
    """Handle damages section creation"""
    write_damages_section(ws, data)


def _handle_additional_sheets(wb, data):
    """Handle CMR and delivery slip images"""
    cmr_url = data.get('cmr_image')
    if cmr_url:
        insert_cmr_sheet(wb.active, cmr_url)

    delivery_slip_images = data.get('delivery_slip_images_urls', [])
    if delivery_slip_images:
        insert_images_in_single_sheet(wb, delivery_slip_images, "Delivery Slips")

    additional_images = data.get('additional_images_urls', [])
    if additional_images:
        insert_images_in_single_sheet(wb, additional_images, "Additional Images")


def _save_final_workbook(wb, relative_path):
    """Save final workbook with all modifications"""
    output_final = None
    try:
        output_final = io.BytesIO()
        wb.save(output_final)
        output_final.seek(0)
        with default_storage.open(relative_path, 'wb') as f:
            f.write(output_final.read())
    finally:
        if output_final:
            output_final.close()


def write_items_to_excel(ws, items, start_row):
    for idx, entry in enumerate(items):
        row = start_row + idx
        ws[get_top_left_cell(ws, f'E{row}')] = "Item"
        ws[get_top_left_cell(ws, f'F{row}')] = entry["item"]["name"]
        ws[get_top_left_cell(ws, f'I{row}')] = "Amount"
        ws[get_top_left_cell(ws, f'J{row}')] = entry["quantity"]


def append_extra_items_to_excel(file_path, extra_items, insert_row):
    wb = None
    try:
        with default_storage.open(file_path, 'rb') as f:
            wb = load_workbook(f)
            ws = wb.active
            affected_merges = []
            for rng in list(ws.merged_cells.ranges):
                if rng.min_row >= insert_row:
                    affected_merges.append((rng.min_row, rng.min_col, rng.max_row, rng.max_col))
                    ws.unmerge_cells(str(rng))
            ws.insert_rows(insert_row, amount=len(extra_items))
            row_above = insert_row - 1
            for i in range(len(extra_items)):
                copy_row_style(ws, row_above, insert_row + i, min_col=1, max_col=12)
            for i in range(len(extra_items)):
                row = insert_row + i
                ws.merge_cells(f"A{row}:B{row}")
                ws.merge_cells(f"C{row}:D{row}")
                ws.merge_cells(f"F{row}:H{row}")
                ws.merge_cells(f"J{row}:L{row}")
                cell = ws.cell(row=row, column=12)
                original = cell.border
                cell.border = Border(
                    left=original.left,
                    right=Side(style="thick"),
                    top=original.top,
                    bottom=original.bottom,
                    diagonal=original.diagonal,
                    diagonal_direction=original.diagonal_direction,
                    outline=original.outline,
                    vertical=original.vertical,
                    horizontal=original.horizontal,
                )
            for min_row, min_col, max_row, max_col in affected_merges:
                new_min_row = min_row + len(extra_items)
                new_max_row = max_row + len(extra_items)
                start_cell = f"{get_column_letter(min_col)}{new_min_row}"
                end_cell = f"{get_column_letter(max_col)}{new_max_row}"
                ws.merge_cells(f"{start_cell}:{end_cell}")
            write_items_to_excel(ws, extra_items, insert_row)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            with default_storage.open(file_path, 'wb') as f:
                f.write(output.read())
    finally:
        if wb:
            wb.close()


def copy_row_style(ws, src_row, dest_row, min_col=1, max_col=12):
    from copy import copy
    for col in range(min_col, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dest_cell = ws.cell(row=dest_row, column=col)
        if src_cell.has_style:
            dest_cell.font = copy(src_cell.font)
            dest_cell.border = copy(src_cell.border)
            dest_cell.fill = copy(src_cell.fill)
            dest_cell.number_format = copy(src_cell.number_format)
            dest_cell.protection = copy(src_cell.protection)
            dest_cell.alignment = copy(src_cell.alignment)


def get_top_left_cell(ws, cell):
    for merged_range in ws.merged_cells.ranges:
        if cell in merged_range:
            return merged_range.start_cell.coordinate
    return cell


def autofit_row_height(ws, cell, text, multiplier=14):
    col_letter = ''.join(filter(str.isalpha, cell))
    row_num = int(''.join(filter(str.isdigit, cell)))
    col_width = ws.column_dimensions[col_letter].width or 8.43
    chars_per_line = int(col_width * 1.15)
    text = (text or "").rstrip()
    total_lines = 0
    for line in text.split('\n'):
        line = line.rstrip()
        if not line:
            total_lines += 1
        else:
            total_lines += (len(line) - 1) // chars_per_line + 1
    ws.row_dimensions[row_num].height = max(15, total_lines * multiplier)


def get_relative_and_abs_path(file_path=None, subdir="delivery_reports_excel", ext="xlsx"):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if file_path:
        # If file_path is provided, use it as relative path
        if isinstance(file_path, Path):
            relative_path = str(file_path)
        else:
            relative_path = file_path
        # For S3, we don't need absolute paths
        abs_path = relative_path
    else:
        # Generate new filename
        relative_path = f"{subdir}/delivery_report_{timestamp}.{ext}"
        abs_path = relative_path

    return relative_path, abs_path


def set_table_outer_border(ws, min_row, max_row, min_col, max_col, border_side):
    for col in range(min_col, max_col + 1):
        # Top edge
        cell = ws.cell(row=min_row, column=col)
        cell.border = Border(top=border_side, left=cell.border.left, right=cell.border.right, bottom=cell.border.bottom)
        # Bottom edge
        cell = ws.cell(row=max_row, column=col)
        cell.border = Border(bottom=border_side, left=cell.border.left, right=cell.border.right, top=cell.border.top)
    for row in range(min_row, max_row + 1):
        # Left edge
        cell = ws.cell(row=row, column=min_col)
        cell.border = Border(left=border_side, top=cell.border.top, right=cell.border.right, bottom=cell.border.bottom)
        # Right edge
        cell = ws.cell(row=row, column=max_col)
        cell.border = Border(right=border_side, top=cell.border.top, left=cell.border.left, bottom=cell.border.bottom)


def write_damages_section(ws, data):
    """
    Writes the Damages section (header, description, images) to the worksheet.
    Applies a thick outer border to the entire damages table.
    """
    damage_description = data.get('damage_description')
    damage_images = data.get('damage_images_urls', [])

    if not (damage_description or damage_images):
        return

    start_row = ws.max_row + 2  # Add some space
    current_row = start_row

    # Styles
    arial_10 = Font(name="Arial", size=10)
    arial_11_bold = Font(name="Arial", size=11, bold=True)
    arial_12_bold = Font(name="Arial", size=12, bold=True)
    header_fill = PatternFill(start_color="8EAADB", end_color="8EAADB", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Damages header (A-L merged, centered, 12pt bold)
    ws.merge_cells(f'A{current_row}:L{current_row}')
    header_cell = ws[f'A{current_row}']
    header_cell.value = "Damages"
    header_cell.font = arial_12_bold
    header_cell.fill = header_fill
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_cell.border = border
    for col in range(1, 13):
        ws[f'{get_column_letter(col)}{current_row}'].font = arial_12_bold
        ws[f'{get_column_letter(col)}{current_row}'].border = border

    # Description row
    current_row += 1
    if damage_description:
        ws[f'A{current_row}'] = "Description:"
        ws[f'A{current_row}'].font = arial_11_bold
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'A{current_row}'].border = border
        ws.merge_cells(f'B{current_row}:L{current_row}')
        desc_cell = ws[f'B{current_row}']
        desc_cell.value = damage_description
        desc_cell.font = arial_10
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
        desc_cell.border = border
        autofit_row_height(ws, f'B{current_row}', damage_description, multiplier=1.5)
        for col in range(2, 13):
            ws[f'{get_column_letter(col)}{current_row}'].font = arial_10
            ws[f'{get_column_letter(col)}{current_row}'].border = border

    # Images section
    if damage_images:
        current_row += 1
        img_start_row = current_row
        img_end_row = current_row + 13

        ws.merge_cells(f'A{img_start_row}:A{img_end_row}')
        img_header_cell = ws[f'A{img_start_row}']
        img_header_cell.value = "Images:"
        img_header_cell.font = arial_11_bold
        img_header_cell.alignment = Alignment(horizontal='center', vertical='center')
        img_header_cell.border = border

        ws.merge_cells(f'B{img_start_row}:L{img_end_row}')
        img_cell = ws[f'B{img_start_row}']
        img_cell.font = arial_10
        img_cell.border = border
        img_cell.alignment = Alignment(vertical='center')

        for row in range(img_start_row, img_end_row + 1):
            for col in range(1, 13):
                cell = ws[f'{get_column_letter(col)}{row}']
                cell.border = border
                if col > 1:
                    cell.font = arial_10
        row_offset = 1 if len(damage_images) > 3 else 7
        create_collage_of_images(
            ws,
            image_urls=[img['image'] if isinstance(img, dict) else img for img in damage_images],
            start_cell=f'B{img_start_row}',
            end_cell=f'L{img_end_row}',
            row_offset=row_offset
        )
        current_row = img_end_row

    end_row = current_row

    # Apply thick outer border to the entire damages table
    outer_side = Side(style="medium")
    set_table_outer_border(ws, min_row=start_row, max_row=end_row, min_col=1, max_col=12, border_side=outer_side)


def insert_client_logo(ws, image_url, cell="I3", end_cell="L7"):
    # Calculate merged cell area in pixels
    start_col = column_index_from_string(''.join(filter(str.isalpha, cell)))
    start_row = int(''.join(filter(str.isdigit, cell)))
    end_col = column_index_from_string(''.join(filter(str.isalpha, end_cell)))
    end_row = int(''.join(filter(str.isdigit, end_cell)))

    total_width = sum(
        int((ws.column_dimensions[get_column_letter(col)].width or 8.43) * 6)
        for col in range(start_col, end_col + 1)
    )
    total_height = sum(
        int((ws.row_dimensions[row].height or 15) * 1.33)
        for row in range(start_row, end_row + 1)
    )

    # Fetch and process image
    img_bytes = BytesIO(fetch_image_bytes(image_url))
    if not img_bytes.getbuffer().nbytes:
        logger.warning(f"Client logo not found or empty: {image_url}")
        return
    pil_img = PILImage.open(img_bytes)
    pil_img = ImageOps.exif_transpose(pil_img)
    pil_img = pil_img.convert("RGBA")
    pil_img.thumbnail((total_width, total_height), PILImage.LANCZOS)

    # Create a transparent canvas and paste the image centered
    canvas = PILImage.new("RGBA", (total_width, total_height), (255, 255, 255, 0))
    x = (total_width - pil_img.width) // 2
    y = (total_height - pil_img.height) // 2 + 3  # Add 3 pixels to avoid cropping
    canvas.paste(pil_img, (x, y), pil_img)

    output = BytesIO()
    canvas.save(output, format="PNG")
    output.seek(0)
    xl_img = XLImage(output)
    xl_img.anchor = cell
    ws.add_image(xl_img)
    # Apply border to the merged range
    border = Side(style="medium")
    for col in range(start_col, end_col + 1):
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(left=cell.border.left, right=border, top=cell.border.top, bottom=cell.border.bottom)
