import io
import math
import requests
import logging
from PIL import Image as PILImage
from PIL import ImageOps
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles import Font, Alignment
import boto3
from urllib.parse import urlparse
from django.conf import settings
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor
import functools
from .file_validators import validate_image_file, FileValidationError

logger = logging.getLogger(__name__)


def create_collage_of_images(ws, image_urls, start_cell, end_cell, row_offset=7):
    n_images = len(image_urls)
    if n_images == 0:
        return
    max_width, max_height = get_range_dimensions(ws, start_cell, add_rows_to_cell(end_cell, row_offset))
    area_aspect_ratio = max_width / max_height
    best_layout = None
    best_size = 0
    for cols in range(1, n_images + 1):
        rows = math.ceil(n_images / cols)
        cell_width = max_width // cols
        cell_height = max_height // rows
        size = min(cell_width, cell_height)
        grid_aspect_ratio = (cols * cell_width) / (rows * cell_height)
        if size > best_size or (size == best_size and abs(grid_aspect_ratio - area_aspect_ratio) < abs(
                (best_layout or (0, 0, 0))[2] - area_aspect_ratio)):
            best_layout = (cols, rows, grid_aspect_ratio)
            best_size = size
    cols, rows, _ = best_layout
    cell_width = max_width // cols
    cell_height = max_height // rows
    last_row_images = n_images % cols if n_images % cols != 0 else cols
    used_width = cell_width * (cols if n_images > cols else n_images) if n_images > (
            rows - 1) * cols else cell_width * last_row_images
    used_height = cell_height * (rows - 1) + (cell_height if last_row_images else 0)
    collage = PILImage.new("RGBA", (cell_width * cols, cell_height * rows), (255, 255, 255, 0))

    # Process images in parallel
    with ThreadPoolExecutor(max_workers=min(4, len(image_urls))) as executor:
        # Create partial function with fixed dimensions
        process_func = functools.partial(fetch_and_process_image,
                                         cell_width=cell_width,
                                         cell_height=cell_height)
        # Submit all image processing tasks
        future_to_index = {
            executor.submit(process_func, url): idx
            for idx, url in enumerate(image_urls)
        }
        # Collect results as they complete
        for future in concurrent.futures.as_completed(future_to_index):
            idx = future_to_index[future]
            try:
                pil_img = future.result(timeout=30)  # 30 second timeout per image
                if pil_img:
                    x = (idx % cols) * cell_width + (cell_width - pil_img.width) // 2
                    y = (idx // cols) * cell_height + (cell_height - pil_img.height) // 2
                    collage.paste(pil_img, (x, y), pil_img)
            except Exception as e:
                logger.error(f"Failed to process image {idx} from {image_urls[idx]}: {e}")

    collage = collage.crop((0, 0, used_width, used_height))
    output_img = transform_image(collage)
    img = XLImage(output_img)
    img.anchor = start_cell
    ws.add_image(img)


def fetch_and_process_image(url, cell_width, cell_height):
    try:
        img_bytes = io.BytesIO(fetch_image_bytes(url))
        if not img_bytes.getbuffer().nbytes:
            return None
            # Validate the fetched image content
            try:
                # Create a mock file object for validation
                img_bytes.seek(0)
                content = img_bytes.read(8192)
                img_bytes.seek(0)

                # Basic validation of image content
                if not _is_valid_image_content(content):
                    logger.warning(f"Invalid image content from {url}")
                    return None

            except Exception as e:
                logger.error(f"Image validation failed for {url}: {e}")
                return None
        pil_img = PILImage.open(img_bytes)
        pil_img = ImageOps.exif_transpose(pil_img)
        pil_img = pil_img.convert("RGBA")

        ratio = min(cell_width / pil_img.width, cell_height / pil_img.height, 1)
        new_size = (int(pil_img.width * ratio), int(pil_img.height * ratio))
        pil_img = pil_img.resize(new_size, PILImage.LANCZOS)

        return pil_img
    except Exception as e:
        logger.error(f"Error processing image from {url}: {e}")
        return None


def insert_images_in_single_sheet(wb, images, sheet_title):
    """
    Insert all images into a single sheet, one below the other.
    """
    img_ws = wb.create_sheet(title=sheet_title)
    img_ws.page_margins.top = 0
    img_ws.page_margins.bottom = 0
    img_ws.page_margins.left = 0.2
    img_ws.page_margins.right = 0.2

    row = 1
    max_page_height = 1060
    descriptor_height = 20
    max_width = 700
    image_height = max_page_height - descriptor_height

    # Process images in parallel first
    processed_images = []
    with ThreadPoolExecutor(max_workers=min(4, len(images))) as executor:
        future_to_index = {}
        for idx, img_obj in enumerate(images):
            url = img_obj.get('image') if isinstance(img_obj, dict) else img_obj
            if url:
                future = executor.submit(process_single_image, url, max_width, image_height)
                future_to_index[future] = idx
        results = {}
        for future in concurrent.futures.as_completed(future_to_index):
            idx = future_to_index[future]
            try:
                result = future.result(timeout=30)
                results[idx] = result
            except Exception as e:
                logger.error(f"Failed to process image {idx}: {e}")
                results[idx] = None

    for idx, img_obj in enumerate(images):
        if idx > 0:
            img_ws.insert_rows(row)
            # Do NOT set row height for the blank row; let Excel use default
            img_ws.row_breaks.append(Break(id=row))
            row += 1

        # Descriptor row
        label = f"{sheet_title[:-1]} {idx + 1}"
        img_ws[f'A{row}'] = label
        img_ws[f'A{row}'].font = Font(bold=True, size=12)
        img_ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        img_ws.row_dimensions[row].height = descriptor_height
        row += 1

        if idx in results and results[idx]:
            xl_img, img_row_height = results[idx]
            cell = f'A{row}'
            xl_img.anchor = cell
            img_ws.add_image(xl_img)
            img_ws.row_dimensions[row].height = img_row_height
            row += 1
    setup_image_worksheet_page(img_ws)


def insert_cmr_sheet(ws, cmr_url=None):
    wb = ws.parent
    max_page_height = 1060
    descriptor_height = 20
    max_width = 700
    image_height = max_page_height - descriptor_height
    if not cmr_url:
        return
    try:
        img_ws = wb.create_sheet(title="CMR")
        img_ws.page_margins.top = 0
        img_ws.page_margins.bottom = 0
        img_ws.page_margins.left = 0.2
        img_ws.page_margins.right = 0.2
        img_ws["A1"] = "CMR Image:"
        img_ws["A1"].font = Font(bold=True, size=12)
        img_ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        img_ws.row_dimensions[1].height = descriptor_height

        # Process image asynchronously (though single image, pattern is consistent)
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(process_single_image, cmr_url, max_width, image_height)
            try:
                result = future.result(timeout=30)
                if result:
                    xl_img, img_row_height = result
                    xl_img.anchor = "A2"
                    img_ws.add_image(xl_img)
                    img_ws.row_dimensions[2].height = img_row_height
            except Exception as e:
                logger.error(f"Failed to process CMR image: {e}")
        setup_image_worksheet_page(img_ws)
    except Exception as e:
        logger.error(f"Error inserting CMR image from {cmr_url}: {e}")


def _is_valid_image_content(content):
    """
    Basic validation of image file content.
    """
    if len(content) < 10:
        return False

    # Check for common image signatures
    image_signatures = [
        b'\xff\xd8\xff',  # JPEG
        b'\x89\x50\x4e\x47',  # PNG
        b'GIF87a', b'GIF89a',  # GIF
        b'BM',  # BMP
        b'RIFF',  # WEBP (partial)
        b'II*\x00', b'MM\x00*'  # TIFF
    ]

    for sig in image_signatures:
        if content.startswith(sig):
            return True

    # Special case for WEBP
    if b'RIFF' in content[:12] and b'WEBP' in content[:12]:
        return True

    return False

def setup_image_worksheet_page(img_ws):
    img_ws.page_setup.orientation = "portrait"
    img_ws.page_setup.paperSize = img_ws.PAPERSIZE_A4
    img_ws.page_setup.fitToPage = False  # Disable fit to page so page breaks work
    img_ws.page_setup.fitToWidth = None
    img_ws.page_setup.fitToHeight = None
    img_ws.page_setup.scale = None


def get_range_dimensions(ws, start_cell, end_cell):
    from openpyxl.utils import get_column_letter, column_index_from_string
    start_col = ''.join(filter(str.isalpha, start_cell))
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    end_col = ''.join(filter(str.isalpha, end_cell))
    end_row = int(''.join(filter(str.isdigit, end_cell)))
    total_width = sum(
        (ws.column_dimensions[get_column_letter(col_idx)].width or 8.43) * 7
        for col_idx in range(column_index_from_string(start_col), column_index_from_string(end_col) + 1)
    )
    total_height = sum(
        (ws.row_dimensions[row].height or 15) * 0.75
        for row in range(start_row, end_row + 1)
    )
    return int(total_width), int(total_height)


def fetch_image_bytes(url):
    """
    Fetch image bytes from S3 if the URL is an S3 object, else use requests.
    """
    try:
        if url.startswith("s3://"):
            s3 = boto3.client("s3")
            parsed = urlparse(url)
            bucket = parsed.netloc
            key = parsed.path.lstrip("/")
            obj = s3.get_object(Bucket=bucket, Key=key)
            content = obj["Body"].read()
        elif settings.AWS_STORAGE_BUCKET_NAME in url:
            s3 = boto3.client("s3")
            parsed = urlparse(url)
            bucket = parsed.netloc.split(".")[0]
            key = parsed.path.lstrip("/")
            obj = s3.get_object(Bucket=bucket, Key=key)
            content = obj["Body"].read()
        else:
            response = requests.get(url, timeout=10, stream=True)
            response.raise_for_status()

            # Check content type
            content_type = response.headers.get('content-type', '').lower()
            if not content_type.startswith('image/'):
                logger.warning(f"Invalid content type for image URL {url}: {content_type}")
                return b''

            content = response.content

        # Validate downloaded content is actually an image
        if not _is_valid_image_content(content):
            logger.warning(f"Downloaded content from {url} is not a valid image")
            return b''

        return content

    except Exception as e:
        logger.error(f"Error fetching image from {url}: {e}")
        return b''

def _is_valid_image_content(content):
    """
    Basic validation of image file content.
    """
    if len(content) < 10:
        return False

    # Check for common image signatures
    image_signatures = [
        b'\xff\xd8\xff',  # JPEG
        b'\x89\x50\x4e\x47',  # PNG
        b'GIF87a', b'GIF89a',  # GIF
        b'BM',  # BMP
        b'RIFF',  # WEBP (partial)
        b'II*\x00', b'MM\x00*'  # TIFF
    ]

    for sig in image_signatures:
        if content.startswith(sig):
            return True

    # Special case for WEBP - check for both RIFF and WEBP signatures
    if b'RIFF' in content[:12] and b'WEBP' in content[:12]:
        return True

    return False

def transform_image(pil_img):
    output_img = io.BytesIO()
    # Save as JPEG if no alpha, else PNG
    if pil_img.mode in ("RGBA", "LA") or (pil_img.mode == "P" and "transparency" in pil_img.info):
        pil_img = pil_img.convert("RGBA")
        pil_img.save(output_img, format="PNG")
    else:
        pil_img = pil_img.convert("RGB")
        pil_img.save(output_img, format="JPEG", quality=85)
    output_img.seek(0)
    return output_img


def add_rows_to_cell(cell, n):
    col = ''.join(filter(str.isalpha, cell))
    row = int(''.join(filter(str.isdigit, cell)))
    return f"{col}{row + n}"


def process_single_image(url, max_width, image_height):
    """
    Process a single image and return XLImage and row height.
    """
    try:
        img_bytes = io.BytesIO(fetch_image_bytes(url))
        pil_img = PILImage.open(img_bytes)
        pil_img = ImageOps.exif_transpose(pil_img)
        pil_img = pil_img.convert("RGBA")

        ratio = min(max_width / pil_img.width, image_height / pil_img.height, 1)
        new_size = (int(pil_img.width * ratio), int(pil_img.height * ratio))
        pil_img = pil_img.resize(new_size, PILImage.LANCZOS)

        output_img = io.BytesIO()
        pil_img.save(output_img, format="PNG")
        output_img.seek(0)

        xl_img = XLImage(output_img)
        img_row_height = new_size[1] * 0.75

        return xl_img, img_row_height
    except Exception as e:
        logger.error(f"Error processing image from {url}: {e}")
        return None
